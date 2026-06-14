from io import BytesIO
from pathlib import Path
from zipfile import ZipFile

from openpyxl import load_workbook
from fastapi.testclient import TestClient
from sqlalchemy import select

from app.database import SessionLocal
from app.main import app
from app.models import AnalysisResult, Call, CallFile, CriterionScore, ManagerProfile, UploadBatch


def auth_headers(client: TestClient) -> dict[str, str]:
    response = client.post("/auth/login", json={"email": "admin@example.com", "password": "admin123"})
    assert response.status_code == 200
    data = response.json()
    return {"Authorization": f"Bearer {data['access_token']}", "X-CSRF-Token": data["csrf_token"]}


def test_upload_process_and_export_call() -> None:
    with TestClient(app) as client:
        headers = auth_headers(client)
        files = [
            ("files", ("2026-06-13_10-35-22_Иванов_79991234567.mp3", b"fake-audio", "audio/mpeg")),
        ]
        response = client.post("/batches", headers=headers, data={"title": "Test batch"}, files=files)
        assert response.status_code == 200, response.text
        batch = response.json()
        assert batch["total_files"] == 1

        calls = client.get("/calls", headers=headers).json()
        assert calls
        detail = client.get(f"/calls/{calls[0]['id']}", headers=headers)
        assert detail.status_code == 200
        payload = detail.json()
        assert payload["transcript"]["segments"]
        assert payload["analysis"]["overall_score"] >= 0

        with SessionLocal() as db:
            analysis = db.scalar(select(AnalysisResult).where(AnalysisResult.call_id == calls[0]["id"]))
            assert analysis is not None
            analysis.evidence = ["plain top-level quote"]
            criterion_score = db.scalar(select(CriterionScore).where(CriterionScore.analysis_result_id == analysis.id))
            assert criterion_score is not None
            criterion_score.evidence = ["plain criterion quote"]
            db.commit()
        detail_with_string_evidence = client.get(f"/calls/{calls[0]['id']}", headers=headers)
        assert detail_with_string_evidence.status_code == 200
        evidence_payload = detail_with_string_evidence.json()
        assert evidence_payload["analysis"]["evidence"][0]["quote"] == "plain top-level quote"
        assert evidence_payload["analysis"]["criteria"][0]["evidence"][0]["quote"] == "plain criterion quote"

        export = client.get("/exports/calls.xlsx", headers=headers)
        assert export.status_code == 200
        assert export.headers["content-type"].startswith("application/vnd.openxmlformats")

        dialogue_export = client.get("/exports/calls-dialogues.xlsx", headers=headers)
        assert dialogue_export.status_code == 200
        workbook = load_workbook(BytesIO(dialogue_export.content))
        sheet = workbook["Calls"]
        headers_row = [cell.value for cell in sheet[1]]
        assert "dialogue" in headers_row

        single_export = client.get(f"/exports/calls/{calls[0]['id']}.xlsx", headers=headers)
        assert single_export.status_code == 200
        single_workbook = load_workbook(BytesIO(single_export.content))
        assert {"Отчет", "Диалог", "Критерии"}.issubset(set(single_workbook.sheetnames))


def test_zip_upload_is_safely_flattened() -> None:
    buffer = BytesIO()
    with ZipFile(buffer, "w") as archive:
        archive.writestr("../evil_good.mp3", b"fake-audio")
    buffer.seek(0)

    with TestClient(app) as client:
        headers = auth_headers(client)
        response = client.post(
            "/batches",
            headers=headers,
            data={"title": "Zip batch"},
            files=[("files", ("calls.zip", buffer.getvalue(), "application/zip"))],
        )
        assert response.status_code == 200, response.text
        assert response.json()["total_files"] == 1


def test_batch_background_processing_drains_queue() -> None:
    with TestClient(app) as client:
        headers = auth_headers(client)
        files = [
            ("files", (f"queue-{index}.mp3", b"fake-audio", "audio/mpeg"))
            for index in range(4)
        ]
        response = client.post("/batches", headers=headers, data={"title": "Queue batch"}, files=files)
        assert response.status_code == 200, response.text
        batch_id = response.json()["id"]

        calls = client.get(f"/calls?batch_id={batch_id}&limit=1000", headers=headers).json()
        assert len(calls) == 4
        assert {call["status"] for call in calls} == {"completed"}


def test_delete_batch_removes_calls_and_files() -> None:
    with TestClient(app) as client:
        headers = auth_headers(client)
        response = client.post(
            "/batches",
            headers=headers,
            data={"title": "Delete me"},
            files=[("files", ("delete-me.mp3", b"fake-audio", "audio/mpeg"))],
        )
        assert response.status_code == 200, response.text
        batch_id = response.json()["id"]

        with SessionLocal() as db:
            call = db.scalar(select(Call).where(Call.batch_id == batch_id))
            assert call is not None
            call_file = db.scalar(select(CallFile).where(CallFile.call_id == call.id))
            assert call_file is not None
            stored_path = Path(call_file.stored_path)
            assert stored_path.exists()

        delete_response = client.delete(f"/batches/{batch_id}", headers=headers)
        assert delete_response.status_code == 200, delete_response.text
        assert delete_response.json()["status"] == "deleted"

        with SessionLocal() as db:
            assert db.get(UploadBatch, batch_id) is None
            assert db.scalar(select(Call).where(Call.batch_id == batch_id)) is None
            assert db.scalar(select(CallFile).where(CallFile.stored_path == str(stored_path))) is None
        assert not stored_path.exists()


def test_manager_comparison_report_and_delete_manager() -> None:
    with TestClient(app) as client:
        headers = auth_headers(client)
        manager_response = client.post(
            "/managers",
            headers=headers,
            json={"name": "Temporary Manager", "department": "Sales", "is_active": True},
        )
        assert manager_response.status_code == 200, manager_response.text
        manager_id = manager_response.json()["id"]
        response = client.post(
            "/batches",
            headers=headers,
            data={"title": "Manager comparison", "manager_id": manager_id},
            files=[("files", ("manager-comparison.mp3", b"fake-audio", "audio/mpeg"))],
        )
        assert response.status_code == 200, response.text

        comparison = client.post("/reports/manager-comparison", headers=headers)
        assert comparison.status_code == 200, comparison.text
        payload = comparison.json()
        assert "manager_rankings" in payload
        assert "service_gaps" in payload
        assert "weaknesses_by_manager" in payload

        delete_response = client.delete(f"/managers/{manager_id}", headers=headers)
        assert delete_response.status_code == 200, delete_response.text
        assert delete_response.json()["status"] == "deleted"
        with SessionLocal() as db:
            assert db.get(ManagerProfile, manager_id) is None
            assert db.scalar(select(Call).where(Call.manager_id == manager_id)) is None


def test_real_audio_under_mock_does_not_get_fake_dialogue() -> None:
    with TestClient(app) as client:
        headers = auth_headers(client)
        response = client.post(
            "/batches",
            headers=headers,
            data={"title": "Real audio under mock"},
            files=[("files", ("25-05-2026__17-25-20__Менеджер 110__74951981216.mp3", b"0" * 256, "audio/mpeg"))],
        )
        assert response.status_code == 200, response.text
        calls = client.get("/calls?limit=1000&q=74951981216", headers=headers).json()
        assert calls
        detail = client.get(f"/calls/{calls[0]['id']}", headers=headers).json()
        assert detail["outcome"] == "no_answer"
        assert detail["overall_score"] == 0
        assert detail["transcript"]["segments"] == []
        assert "Металл Сервис" not in detail["transcript"]["text"]


def test_rbac_requires_auth() -> None:
    with TestClient(app) as client:
        response = client.get("/dashboard")
        assert response.status_code == 401
