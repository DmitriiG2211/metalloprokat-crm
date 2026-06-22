from __future__ import annotations

from io import BytesIO
from typing import Any
import re

import pandas as pd
from fastapi import HTTPException, UploadFile
from openpyxl import load_workbook
from openpyxl.styles.colors import COLOR_INDEX
from sqlalchemy import or_, select
from sqlalchemy.orm import Session

from app.models import Client, ClientComment, ImportClientChange, ImportError, ImportJob, Status, User
from app.services.audit import write_audit
from app.utils.normalization import normalize_company, normalize_email, normalize_phone, normalize_website, parse_date


FIELD_ALIASES = {
    "company_name": [
        "компания",
        "название компании",
        "название клиента",
        "организация",
        "наименование",
        "наименование клиента",
        "наименование контрагента",
        "наименование покупателя",
        "клиент",
        "контрагент",
        "покупатель",
        "заказчик",
        "наименование организации",
        "company",
        "company name",
        "organization",
        "client",
    ],
    "contact_person": ["фио", "контактное лицо", "представитель", "имя", "фио клиента"],
    "phone": ["телефон", "номер телефона", "мобильный", "контактный телефон", "phone", "phone number", "tel"],
    "email": ["почта", "email", "e-mail", "электронная почта", "mail", "эл. почта"],
    "website": ["сайт", "web", "website", "url", "ссылка", "ссылка на сайт", "адрес сайта", "интернет", "site"],
    "comment": ["комментарий", "итог звонка", "примечание", "заметка", "результат", "comment", "note", "result"],
    "last_call_date": ["дата звонка", "последний звонок", "когда звонили", "дата первичного звонка", "дата повторного звонка"],
    "next_call_date": ["дата перезвона", "перезвонить", "следующий звонок", "дата следующего звонка"],
}

GENERIC_LINK_TEXT = {"сайт", "link", "url", "website", "web", "перейти", "открыть"}


def _normalize_header(value: object) -> str:
    text = str(value or "").strip().lower()
    text = text.replace("ё", "е")
    return re.sub(r"[\s_./\\\-:;№#()]+", " ", text).strip()


def _looks_like_email(value: object) -> bool:
    return "@" in str(value or "")


def _looks_like_website(value: object) -> bool:
    text = str(value or "").strip().lower()
    if not text:
        return False
    return text.startswith(("http://", "https://", "www.")) or bool(re.search(r"\b[a-zа-я0-9-]+\.(ru|рф|com|net|org|su|biz|info)\b", text))


def _looks_like_phone(value: object) -> bool:
    return len(normalize_phone(str(value or "")) or "") >= 7


def _looks_like_date(value: object) -> bool:
    return parse_date(value) is not None


def _looks_like_company(value: object) -> bool:
    text = str(value or "").strip()
    if len(text) < 3:
        return False
    if _looks_like_phone(text) or _looks_like_email(text) or _looks_like_website(text) or _looks_like_date(text):
        return False
    if re.fullmatch(r"[\d\s.,+-]+", text):
        return False
    company_markers = ("ооо", "ип ", "ао ", "оао", "зао", "пао", "пкф", "тд ", "тк ", "металл", "строй", "снаб", "производ")
    lower = text.lower()
    return any(marker in lower for marker in company_markers) or bool(re.search(r"[а-яА-Яa-zA-Z]{4,}", text))


def _read_dataframe(raw: bytes, name: str) -> pd.DataFrame:
    try:
        if name.endswith(".csv"):
            try:
                df = pd.read_csv(BytesIO(raw), dtype=object)
            except UnicodeDecodeError:
                df = pd.read_csv(BytesIO(raw), dtype=object, encoding="cp1251")
            df.columns = [str(column) for column in df.columns]
            return df
        if name.endswith((".xlsx", ".xls")):
            df = pd.read_excel(BytesIO(raw), dtype=object)
            df.columns = [str(column) for column in df.columns]
            return df
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Не удалось прочитать файл: {exc}") from exc
    raise HTTPException(status_code=400, detail="Поддерживаются только .xlsx, .xls и .csv")


def _cell_rgb(cell) -> str | None:
    fill = cell.fill
    color = fill.fgColor if fill and fill.fill_type else None
    if not color:
        return None
    if color.type == "rgb" and color.rgb:
        rgb = str(color.rgb)[-6:].upper()
    elif color.type == "indexed" and color.indexed is not None and color.indexed < len(COLOR_INDEX):
        rgb = str(COLOR_INDEX[color.indexed])[-6:].upper()
    else:
        return None
    if rgb in {"000000", "FFFFFF"}:
        return None
    return rgb


def _status_from_rgb(rgb: str | None) -> str | None:
    if not rgb or len(rgb) != 6:
        return None
    red = int(rgb[0:2], 16)
    green = int(rgb[2:4], 16)
    blue = int(rgb[4:6], 16)
    if red >= 220 and green <= 215 and blue <= 220:
        return "Мертвый"
    if red >= 220 and green >= 180 and blue <= 210:
        return "50/50"
    if green >= 150 and green >= red + 15 and green >= blue + 15:
        return "Контактный"
    return None


def _detect_row_statuses(raw: bytes, name: str) -> dict[int, str]:
    if not name.endswith((".xlsx", ".xlsm")):
        return {}
    statuses: dict[int, str] = {}
    try:
        workbook = load_workbook(BytesIO(raw), data_only=True, read_only=False)
    except Exception:
        return statuses
    sheet = workbook.active
    priority = {"Мертвый": 3, "50/50": 2, "Контактный": 1}
    for row in sheet.iter_rows(min_row=2):
        row_scores: dict[str, int] = {}
        for cell in row:
            status = _status_from_rgb(_cell_rgb(cell))
            if status:
                row_scores[status] = row_scores.get(status, 0) + 1
        if row_scores:
            statuses[row[0].row] = max(row_scores, key=lambda status: (row_scores[status], priority[status]))
    return statuses


def _detect_cell_hyperlinks(raw: bytes, name: str) -> dict[int, dict[int, str]]:
    if not name.endswith((".xlsx", ".xlsm")):
        return {}
    links: dict[int, dict[int, str]] = {}
    try:
        workbook = load_workbook(BytesIO(raw), data_only=True, read_only=False)
    except Exception:
        return links
    sheet = workbook.active
    for row in sheet.iter_rows(min_row=2):
        for zero_based_index, cell in enumerate(row):
            target = getattr(cell.hyperlink, "target", None) if cell.hyperlink else None
            if target:
                links.setdefault(cell.row, {})[zero_based_index] = str(target).strip()
    return links


async def read_upload(file: UploadFile) -> pd.DataFrame:
    raw = await file.read()
    name = (file.filename or "").lower()
    return _read_dataframe(raw, name)


async def read_upload_with_row_statuses(file: UploadFile) -> tuple[pd.DataFrame, dict[int, str], dict[int, dict[int, str]]]:
    raw = await file.read()
    name = (file.filename or "").lower()
    return _read_dataframe(raw, name), _detect_row_statuses(raw, name), _detect_cell_hyperlinks(raw, name)


def detect_mapping(columns: list[str]) -> dict[str, str | None]:
    normalized_columns = {_normalize_header(column): column for column in columns}
    mapping: dict[str, str | None] = {}
    for field, aliases in FIELD_ALIASES.items():
        found = None
        normalized_aliases = [_normalize_header(alias) for alias in aliases]
        for alias in aliases:
            normalized_alias = _normalize_header(alias)
            if normalized_alias in normalized_columns:
                found = normalized_columns[normalized_alias]
                break
        if not found:
            for normalized, original in normalized_columns.items():
                if any(alias and alias in normalized for alias in normalized_aliases):
                    found = original
                    break
        mapping[field] = found
    return mapping


def _sample_values(df: pd.DataFrame, column: str) -> list[object]:
    return [value for value in df[column].head(80).tolist() if not pd.isna(value)]


def _infer_column_by_values(df: pd.DataFrame, field: str, used: set[str]) -> str | None:
    best_column: str | None = None
    best_score = 0.0
    for position, column in enumerate(df.columns):
        column_name = str(column)
        if column_name in used:
            continue
        values = _sample_values(df, column_name)
        if not values:
            continue
        checks = {
            "phone": _looks_like_phone,
            "email": _looks_like_email,
            "website": _looks_like_website,
            "last_call_date": _looks_like_date,
            "next_call_date": _looks_like_date,
            "company_name": _looks_like_company,
        }
        check = checks.get(field)
        if not check:
            continue
        matched = sum(1 for value in values if check(value))
        score = matched / max(1, len(values))
        if field == "company_name":
            header = _normalize_header(column_name)
            if any(alias in header for alias in ("клиент", "компания", "контрагент", "организация", "наименование")):
                score += 0.55
            if position == 1:
                score += 0.1
            if any(_looks_like_phone(value) or _looks_like_email(value) or _looks_like_website(value) for value in values):
                score -= 0.35
        if field == "website" and any(_looks_like_website(value) for value in values):
            score += 0.25
        if score > best_score:
            best_score = score
            best_column = column_name
    threshold = 0.15 if field == "company_name" else 0.2
    return best_column if best_score >= threshold else None


def detect_mapping_from_dataframe(df: pd.DataFrame) -> dict[str, str | None]:
    mapping = detect_mapping([str(column) for column in df.columns])
    used = {column for column in mapping.values() if column}
    for field in ("phone", "email", "website", "company_name"):
        if not mapping.get(field):
            inferred = _infer_column_by_values(df, field, used)
            if inferred:
                mapping[field] = inferred
                used.add(inferred)
    return mapping


def preview_dataframe(df: pd.DataFrame, filename: str) -> dict[str, Any]:
    df = df.dropna(how="all")
    return {
        "filename": filename,
        "columns": [str(column) for column in df.columns],
        "mapping": detect_mapping_from_dataframe(df),
        "preview_rows": df.head(20).where(pd.notnull(df), None).to_dict(orient="records"),
        "total_rows": int(len(df)),
    }


def _get_value(row: pd.Series, mapping: dict[str, str | None], field: str):
    column = mapping.get(field)
    if not column or column not in row:
        return None
    value = row[column]
    if pd.isna(value):
        return None
    return value


def _get_hyperlink(row: pd.Series, row_number: int, mapping: dict[str, str | None], field: str, hyperlinks: dict[int, dict[int, str]]) -> str | None:
    column = mapping.get(field)
    if not column or column not in row.index:
        return None
    try:
        column_index = list(row.index).index(column)
    except ValueError:
        return None
    return hyperlinks.get(row_number, {}).get(column_index)


def find_duplicate(db: Session, company_name: str, phone: str | None, email: str | None, domain: str | None) -> Client | None:
    normalized_company = normalize_company(company_name)
    conditions = [Client.normalized_company_name == normalized_company]
    if phone:
        conditions.append(Client.normalized_phone == phone)
    if email:
        first_email = email.splitlines()[0].lower()
        conditions.append(Client.email.ilike(f"%{first_email}%"))
    if domain:
        conditions.append(Client.website_domain == domain)
    return db.scalars(select(Client).where(Client.deleted_at.is_(None), or_(*conditions))).first()


def _safe_raw_data(row: pd.Series) -> dict[str, str | None]:
    raw_data: dict[str, str | None] = {}
    for key, value in row.to_dict().items():
        raw_data[str(key)] = None if pd.isna(value) else str(value)
    return raw_data


def _existing_client_keys(db: Session) -> tuple[set[str], set[str], set[str], set[str]]:
    rows = db.execute(
        select(Client.normalized_company_name, Client.normalized_phone, Client.email, Client.website_domain).where(Client.deleted_at.is_(None))
    ).all()
    companies: set[str] = set()
    phones: set[str] = set()
    emails: set[str] = set()
    domains: set[str] = set()
    for company, phone, email, domain in rows:
        if company:
            companies.add(company)
        if phone:
            phones.add(phone)
        if email:
            first_email = str(email).splitlines()[0].lower()
            if first_email:
                emails.add(first_email)
        if domain:
            domains.add(domain)
    return companies, phones, emails, domains


def import_dataframe(
    db: Session,
    df: pd.DataFrame,
    filename: str,
    uploaded_by: User,
    assigned_manager_id: int,
    mapping: dict[str, str | None] | None = None,
    row_statuses: dict[int, str] | None = None,
    cell_hyperlinks: dict[int, dict[int, str]] | None = None,
) -> ImportJob:
    detected_mapping = detect_mapping_from_dataframe(df)
    supplied_mapping = {key: (value or None) for key, value in (mapping or {}).items()}
    mapping = {**detected_mapping, **{key: value for key, value in supplied_mapping.items() if value}}
    default_status = db.scalars(select(Status).where(Status.is_active.is_(True)).order_by(Status.sort_order)).first()
    status_ids = {status.name.casefold(): status.id for status in db.scalars(select(Status).where(Status.is_active.is_(True))).all()}
    row_statuses = row_statuses or {}
    cell_hyperlinks = cell_hyperlinks or {}
    job = ImportJob(filename=filename, uploaded_by=uploaded_by.id, assigned_manager_id=assigned_manager_id, status="running")
    db.add(job)
    db.flush()

    df = df.dropna(how="all")
    job.total_rows = int(len(df))
    existing_companies, existing_phones, existing_emails, existing_domains = _existing_client_keys(db)
    pending_companies: set[str] = set()
    pending_phones: set[str] = set()
    pending_emails: set[str] = set()
    pending_domains: set[str] = set()

    for index, row in df.iterrows():
        row_number = int(index) + 2
        company = _get_value(row, mapping, "company_name")
        phone_raw = _get_value(row, mapping, "phone")
        if not company and not phone_raw:
            job.skipped_count += 1
            db.add(ImportError(import_id=job.id, row_number=row_number, error_text="Нет названия компании и телефона", raw_data=_safe_raw_data(row)))
            continue

        company_name = str(company or "Без названия").strip()
        normalized_company = normalize_company(company_name)
        phone = str(phone_raw).strip() if phone_raw else None
        normalized_phone = normalize_phone(phone)
        email = normalize_email(_get_value(row, mapping, "email"))
        first_email = email.splitlines()[0].lower() if email else None
        website_raw = _get_value(row, mapping, "website")
        website_link = _get_hyperlink(row, row_number, mapping, "website", cell_hyperlinks)
        if website_link and (not website_raw or str(website_raw).strip().lower() in GENERIC_LINK_TEXT or not _looks_like_website(website_raw)):
            website_raw = website_link
        website, domain = normalize_website(website_raw)
        duplicate = (
            normalized_company in existing_companies
            or normalized_company in pending_companies
            or (normalized_phone and (normalized_phone in existing_phones or normalized_phone in pending_phones))
            or (first_email and (first_email in existing_emails or first_email in pending_emails))
            or (domain and (domain in existing_domains or domain in pending_domains))
        )
        if duplicate:
            job.duplicate_count += 1
        row_status_name = row_statuses.get(row_number)
        row_status_id = status_ids.get(row_status_name.casefold()) if row_status_name else None

        client = Client(
            manager_id=assigned_manager_id,
            company_name=company_name,
            normalized_company_name=normalized_company,
            contact_person=str(_get_value(row, mapping, "contact_person") or "").strip() or None,
            phone=phone,
            normalized_phone=normalized_phone,
            email=email,
            website=website,
            website_domain=domain,
            status_id=row_status_id or (default_status.id if default_status else None),
            last_call_date=parse_date(_get_value(row, mapping, "last_call_date")),
            next_call_date=parse_date(_get_value(row, mapping, "next_call_date")),
            source_import_id=job.id,
        )
        db.add(client)
        db.flush()
        db.add(
            ImportClientChange(
                import_id=job.id,
                client_id=client.id,
                action="created",
                new_data={
                    "company_name": client.company_name,
                    "phone": client.phone,
                    "email": client.email,
                    "website": client.website,
                    "status_id": client.status_id,
                },
            )
        )
        pending_companies.add(normalized_company)
        if normalized_phone:
            pending_phones.add(normalized_phone)
        if first_email:
            pending_emails.add(first_email)
        if domain:
            pending_domains.add(domain)

        comment = _get_value(row, mapping, "comment")
        if comment:
            db.add(ClientComment(client=client, author_id=uploaded_by.id, comment_text=str(comment).strip()))
        job.created_count += 1

    job.status = "done"
    job.error_count = len(db.scalars(select(ImportError).where(ImportError.import_id == job.id)).all())
    write_audit(db, uploaded_by, "import_excel", "import", job.id, new_value={"filename": filename, "created": job.created_count})
    return job
